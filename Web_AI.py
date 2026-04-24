#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import re
import json
import time
import datetime
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

try:
    import requests
except ImportError:
    print("Failed to import requests module. Please install it first.")
    raise


# -----------------------------------------------------------------------------
# LOAD CONFIG
# -----------------------------------------------------------------------------
def load_config(path):
    try:
        with open(path, "r") as f:
            return json.load(f)
    except Exception as e:
        print("Failed to load config: {0}".format(e))
        exit(1)


# -----------------------------------------------------------------------------
# SAFE SSH EXECUTION
# -----------------------------------------------------------------------------
def ssh_exec(host, user, key_path, command):
    cmd = ["ssh", "-q", "-o", "StrictHostKeyChecking=no", "-o", "ConnectTimeout=10"]

    if key_path:
        cmd += ["-i", os.path.expanduser(key_path)]

    cmd.append("{0}@{1}".format(user, host))
    cmd.append(command)

    try:
        res = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        stdout, stderr = res.communicate()

        if not isinstance(stdout, str):
            stdout = stdout.decode("utf-8", "ignore")
        if not isinstance(stderr, str):
            stderr = stderr.decode("utf-8", "ignore")

        return {
            "success": res.returncode in (0, 1),
            "stdout": stdout.strip(),
            "stderr": stderr.strip(),
            "returncode": res.returncode,
        }
    except Exception as e:
        return {"success": False, "stdout": "", "stderr": str(e), "returncode": 1}


# -----------------------------------------------------------------------------
# AI CALL
# -----------------------------------------------------------------------------
def call_ai(prompt):
    api_url = os.getenv("API_URL")
    api_key = os.getenv("API_KEY")

    if not api_url or not api_key:
        return {"error": "AI API not configured."}

    headers = {
        "Authorization": "Bearer {0}".format(api_key),
        "Content-Type": "application/json",
    }

    payload = {
        "model": "gpt-4",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.2,
        "max_tokens": 1100,
    }

    try:
        r = requests.post(api_url, json=payload, headers=headers, timeout=60)
        r.raise_for_status()
        content = r.json()["choices"][0]["message"]["content"].strip()

        # 1. direct JSON
        try:
            return json.loads(content)
        except Exception:
            pass

        # 2. markdown fenced JSON block
        fenced = re.search(r"", content, re.S)
        if fenced:
            try:
                return json.loads(fenced.group(1))
            except Exception:
                pass

        # 3. any JSON object inside mixed text
        generic = re.search(r"(\{.*\})", content, re.S)
        if generic:
            try:
                return json.loads(generic.group(1))
            except Exception:
                pass

        return {"diagnosis": content}

    except Exception as e:
        return {"error": "AI call failed: {0}".format(str(e))}


# -----------------------------------------------------------------------------
# PARSE STATUS
# -----------------------------------------------------------------------------
def parse_websphere_status(output):
    if not output:
        return [], [], []

    configured = re.findall(r"Server name:\s+(\S+)", output)
    started = re.findall(r'"([^"]+)" is STARTED', output)
    down = [srv for srv in configured if srv not in started]

    return configured, started, down


# -----------------------------------------------------------------------------
# GET SHUTDOWN CONTEXT
# -----------------------------------------------------------------------------
def get_latest_shutdown_context(host, profile, jvm, log_path):
    find_cmd = "grep -n 'WSVR0023I: Server {0} is stopping' {1} | tail -1".format(
        jvm, log_path
    )

    result = ssh_exec(host, profile["ssh_user"], profile.get("ssh_key_path"), find_cmd)

    if not result["success"] or not result["stdout"]:
        return None, None

    try:
        line_number = int(result["stdout"].split(":")[0])
    except Exception:
        return None, None

    ts_cmd = "sed -n '{0}p' {1}".format(line_number, log_path)
    ts_res = ssh_exec(host, profile["ssh_user"], profile.get("ssh_key_path"), ts_cmd)

    timestamp_line = ts_res["stdout"] if ts_res["success"] else "UNKNOWN"

    start_line = max(1, line_number - 200)
    end_line = line_number + 50

    ctx_cmd = "sed -n '{0},{1}p' {2}".format(start_line, end_line, log_path)
    ctx_res = ssh_exec(host, profile["ssh_user"], profile.get("ssh_key_path"), ctx_cmd)

    if not ctx_res["success"]:
        return timestamp_line, None

    return timestamp_line, ctx_res["stdout"]


# -----------------------------------------------------------------------------
# NATIVE STDOUT (initial down-JVM triage, after SystemOut)
# -----------------------------------------------------------------------------
def get_native_stdout_tail(host, profile, system_out_log_path, lines=200):
    """
    Tail native_stdout.log in the same directory as SystemOut.log.
    Returns stripped text or empty string if missing/unreadable.
    """
    path = os.path.join(os.path.dirname(system_out_log_path), "native_stdout.log")
    res = ssh_exec(
        host,
        profile["ssh_user"],
        profile.get("ssh_key_path"),
        "tail -n {0} {1}".format(lines, path),
    )
    if not res["success"]:
        return ""
    return (res.get("stdout") or "").rstrip()


# -----------------------------------------------------------------------------
# STARTSERVER.LOG CONFIDENCE (whether native_stderr is needed)
# -----------------------------------------------------------------------------
def is_startserver_log_confident(text):
    """
    True if startServer.log tail is plausibly actionable for root-cause analysis.
    If False, native_stderr.log is tailed as a fallback.
    """
    if not text or not str(text).strip():
        return False
    t = str(text).strip()
    if len(t) < 100:
        return False
    lines = [ln for ln in t.splitlines() if ln.strip()]
    if len(lines) < 2:
        return False
    if len(t) > 900:
        return True
    if re.search(
        r"(?i)(exception|error|severe|failed|fatal|cannot launch|could not|"
        r"unavailable|java\.lang\.|adm\d{4}|wsvr0|admu)",
        t,
    ):
        return True
    return len(lines) >= 14


# -----------------------------------------------------------------------------
# GET STARTUP FAILURE LOGS (after first failed start: startServer, then stderr)
# -----------------------------------------------------------------------------
def get_startup_failure_context(host, profile, system_out_log_path):
    """
    After start attempt 1 fails: read startServer.log first (last 300 lines).
    If that is missing, unreadable, or not diagnostically confident, also tail
    native_stderr.log (last 300 lines). Same server log directory as SystemOut.

    Returns None only if neither startServer nor native_stderr could be read.
    """
    user = profile["ssh_user"]
    key = profile.get("ssh_key_path")
    log_dir = os.path.dirname(system_out_log_path)
    start_server_path = os.path.join(log_dir, "startServer.log")
    native_stderr_path = os.path.join(log_dir, "native_stderr.log")

    ss_res = ssh_exec(
        host, user, key, "tail -n 300 {0}".format(start_server_path)
    )
    ss_text = (ss_res.get("stdout") or "").rstrip() if ss_res["success"] else ""
    parts = []

    if ss_res["success"]:
        parts.append(
            "=== startServer.log (last 300 lines; primary after first failed start) ===\n{0}".format(
                ss_text if ss_text else "(empty)"
            )
        )
    else:
        parts.append(
            "=== startServer.log (last 300 lines) ===\n[!] {0}".format(
                (ss_res.get("stderr") or "tail failed").strip()
            )
        )

    need_stderr = (not ss_res["success"]) or (
        not is_startserver_log_confident(ss_text)
    )
    es_success = False
    if need_stderr:
        es_res = ssh_exec(
            host, user, key, "tail -n 300 {0}".format(native_stderr_path)
        )
        es_success = bool(es_res["success"])
        tag = (
            "included because startServer.log lacked clear failure signals"
            if ss_res["success"]
            else "included because startServer.log could not be read"
        )
        if es_res["success"]:
            block = (es_res.get("stdout") or "").rstrip()
            parts.append(
                "=== native_stderr.log (last 300 lines; {0}) ===\n{1}".format(
                    tag, block if block else "(empty)"
                )
            )
        else:
            parts.append(
                "=== native_stderr.log (last 300 lines; {0}) ===\n[!] {1}".format(
                    tag, (es_res.get("stderr") or "tail failed").strip()
                )
            )

    if ss_res["success"] or es_success:
        return "\n\n".join(parts)
    return None


# -----------------------------------------------------------------------------
# EXCEL REPORT
# -----------------------------------------------------------------------------
REPORT_COLUMNS = ["Patched server", "status", "cause", "ai action"]


def _new_jvm_report():
    return {
        "shutdown_ai": None,
        "startup_ai": None,
        "notes": [],
        "restart_attempted": False,
        "recovered": False,
    }


def _report_row(host, jvm, status, cause, ai_action):
    return {
        "Patched server": "{0} / {1}".format(host, jvm),
        "status": status,
        "cause": cause,
        "ai action": ai_action,
    }


def _jvm_ai_text(info):
    if not info:
        return ""
    parts = []
    if info.get("shutdown_ai"):
        parts.append("Shutdown analysis:\n{0}".format(info["shutdown_ai"]))
    if info.get("startup_ai"):
        parts.append("Startup failure analysis:\n{0}".format(info["startup_ai"]))
    return "\n\n".join(parts)


def _jvm_cause_text(info, running):
    info = info or {}
    if running:
        if info.get("recovered"):
            return "Recovered after automated restart"
        return ""
    parts = []
    if info.get("restart_attempted") and not info.get("recovered"):
        parts.append("Still down after restart attempt(s)")
    parts.extend(info.get("notes") or [])
    return "; ".join(parts) if parts else "Down"


def write_excel_report(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Servers"
    ws.append(REPORT_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in REPORT_COLUMNS])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for idx in range(1, len(REPORT_COLUMNS) + 1):
        letter = get_column_letter(idx)
        col_cells = [ws.cell(row=r, column=idx) for r in range(1, ws.max_row + 1)]
        maxlen = min(max(len(str(c.value or "")) for c in col_cells), 60)
        ws.column_dimensions[letter].width = max(12, maxlen + 2)

    wb.save(output_path)
    return True


def get_generic_l1_context(host, profile, was_home, jvm, system_out_log_path):
    """
    Collect generic L1 diagnostics to help AI determine the right fix
    based on the actual startup issue.
    """
    user = profile["ssh_user"]
    key = profile.get("ssh_key_path")

    system_err_log_path = os.path.join(
        os.path.dirname(system_out_log_path),
        "SystemErr.log",
    )
    native_stderr_log_path = os.path.join(
        os.path.dirname(system_out_log_path),
        "native_stderr.log",
    )
    native_stdout_log_path = os.path.join(
        os.path.dirname(system_out_log_path),
        "native_stdout.log",
    )

    server_xml_glob = "{0}/config/cells/*/nodes/*/servers/{1}/server.xml".format(
        was_home, jvm
    )

    commands = [
        (
            "Memory Summary",
            r"""sh -c 'echo "--- free -m ---"; free -m 2>/dev/null; echo; echo "--- /proc/meminfo (top) ---"; cat /proc/meminfo 2>/dev/null | head -20'""",
        ),
        ("Disk", r"df -h 2>/dev/null"),
        ("Ulimits", r"ulimit -a 2>/dev/null"),
        ("Mounts", r"mount 2>/dev/null | head -50"),
        ("Identity", r"id 2>/dev/null"),
        ("Current Directory", r"pwd 2>/dev/null"),
        (
            "Server XML Path",
            r"""sh -c 'ls {0} 2>/dev/null | head -1'""".format(server_xml_glob),
        ),
        (
            "Heap details from server.xml",
            r"""sh -c 'f=$(ls {0} 2>/dev/null | head -1); if [ -n "$f" ]; then echo "FILE=$f"; grep -i "heap" "$f"; grep -i "jvmEntries" "$f"; else echo "server.xml not found"; fi'""".format(
                server_xml_glob
            ),
        ),
        (
            "genericJvmArguments from server.xml",
            r"""sh -c 'f=$(ls {0} 2>/dev/null | head -1); if [ -n "$f" ]; then echo "FILE=$f"; grep -i "genericJvmArguments" "$f" | head -3; else echo "server.xml not found"; fi'""".format(
                server_xml_glob
            ),
        ),
        ("SystemOut.log", "tail -n 200 {0}".format(system_out_log_path)),
        ("native_stdout.log", "tail -n 200 {0}".format(native_stdout_log_path)),
        ("SystemErr.log", "tail -n 200 {0}".format(system_err_log_path)),
        ("native_stderr.log", "tail -n 200 {0}".format(native_stderr_log_path)),
    ]

    parts = []
    for title, cmd in commands:
        res = ssh_exec(host, user, key, cmd)
        if res["success"]:
            content = res["stdout"].strip() or "(empty)"
        else:
            content = "[!] {0}".format((res.get("stderr") or "command failed").strip())
        parts.append("=== {0} ===\n{1}".format(title, content))

    return "\n\n".join(parts)


# -----------------------------------------------------------------------------
# EXECUTE SAFE L1 VALIDATIONS
# -----------------------------------------------------------------------------
def execute_l1_validations(host, profile, validation_commands):
    """
    Execute only safe AI-generated read-only validation commands.
    """
    user = profile["ssh_user"]
    key = profile.get("ssh_key_path")

    safe_prefixes = (
        "test ",
        "ls ",
        "stat ",
        "file ",
        "netstat ",
        "ss ",
        "echo ",
        "grep ",
        "cat ",
        "head ",
        "tail ",
        "find ",
        "df ",
        "free ",
        "ulimit ",
        "ps ",
        "pwd ",
        "id ",
        "mount ",
    )

    print("\nExecuting AI-generated L1 validation commands:\n")

    outputs = []
    validation_confirmed = False

    for cmd in validation_commands:
        cmd = (cmd or "").strip()

        if not cmd.startswith(safe_prefixes):
            print("  [SKIPPED unsafe] {0}".format(cmd))
            outputs.append(
                {"command": cmd, "executed": False, "output": "Skipped unsafe command"}
            )
            continue

        res = ssh_exec(host, user, key, cmd)

        out = (res.get("stdout") or "").strip()
        err = (res.get("stderr") or "").strip()
        combined = out if out else err

        outputs.append(
            {"command": cmd, "executed": True, "output": combined or "(empty)"}
        )

        text = (combined or "").lower()
        if (
            "not found" in text
            or "no such file" in text
            or "cannot access" in text
            or "permission denied" in text
            or "too large" in text
            or "address already in use" in text
            or "manifest missing" in text
            or "failed" in text
            or "exists" in text
            or "found" in text
        ):
            validation_confirmed = True

    return {"confirmed": validation_confirmed, "outputs": outputs}


# -----------------------------------------------------------------------------
# FORMAT AI DIAGNOSIS
# -----------------------------------------------------------------------------
def format_ai_diagnosis(diag):
    if not isinstance(diag, dict):
        return str(diag)

    if diag.get("error"):
        return diag["error"]

    if diag.get("diagnosis") and len(diag.keys()) == 1:
        return diag["diagnosis"]

    parts = []

    jvm_name = diag.get("jvm_that_is_down") or diag.get("jvm")
    if jvm_name:
        parts.append("1. JVM that is down:\n{0}".format(jvm_name))

    if diag.get("exact_error"):
        parts.append("2. Exact Error:\n{0}".format(diag["exact_error"]))

    if diag.get("root_cause"):
        parts.append("3. Root Cause:\n{0}".format(diag["root_cause"]))

    if diag.get("evidence"):
        parts.append("4. Evidence/L1 diagnostic:\n{0}".format(diag["evidence"]))

    if diag.get("explanation"):
        parts.append("5. Clear Explanation:\n{0}".format(diag["explanation"]))

    if diag.get("recommended_fix"):
        parts.append("6. Recommended Fix:\n{0}".format(diag["recommended_fix"]))

    if diag.get("console_fix_navigation"):
        parts.append(
            "7. Check/Fix from Console by navigating to:\n{0}".format(
                diag["console_fix_navigation"]
            )
        )

    if diag.get("server_start_command"):
        parts.append(
            "8. Command for starting from server:\n{0}".format(
                diag["server_start_command"]
            )
        )

    if diag.get("console_start_navigation"):
        parts.append(
            "9. Start from Console by navigating to:\n{0}".format(
                diag["console_start_navigation"]
            )
        )

    if diag.get("post_validation_recommendation"):
        parts.append(
            "10. Post-validation Recommendation:\n{0}".format(
                diag["post_validation_recommendation"]
            )
        )

    return "\n\n".join(parts) if parts else str(diag)


# -----------------------------------------------------------------------------
# FORMAT L1 VALIDATION OUTPUTS
# -----------------------------------------------------------------------------
def format_validation_outputs(validation_result, profile=None, host=None):
    if not validation_result:
        return ""

    shell_user = "user"
    if profile and profile.get("ssh_user"):
        shell_user = profile.get("ssh_user")

    shell_host = host or "host"

    lines = []
    lines.append("Results in server:")

    for item in validation_result.get("outputs", []):
        lines.append(
            "[{0}@{1} ~]$ {2}".format(shell_user, shell_host, item.get("command", ""))
        )
        lines.append("{0}".format(item.get("output", "(empty)")))
        lines.append("")

    lines.append(
        "Validation confirmed: {0}".format(
            "Yes" if validation_result.get("confirmed") else "No"
        )
    )

    return "\n".join(lines)


def get_exact_server_xml_and_heap_details(host, profile, was_home, jvm):
    """
    Resolve exact server.xml path for the JVM and fetch heap/jvmEntries details.
    """
    user = profile["ssh_user"]
    key = profile.get("ssh_key_path")

    server_xml_glob = "{0}/config/cells/*/nodes/*/servers/{1}/server.xml".format(
        was_home, jvm
    )

    find_cmd = r"""sh -c 'ls {0} 2>/dev/null | head -1'""".format(server_xml_glob)
    find_res = ssh_exec(host, user, key, find_cmd)

    server_xml_path = ""
    if find_res["success"] and find_res["stdout"]:
        server_xml_path = find_res["stdout"].strip()

    if not server_xml_path:
        return {"server_xml_path": "", "heap_lines": "", "jvm_entries_line": ""}

    heap_cmd = r"""sh -c 'grep -i "heap" "{0}"'""".format(server_xml_path)
    heap_res = ssh_exec(host, user, key, heap_cmd)

    jvm_cmd = r"""sh -c 'grep -i "jvmEntries" "{0}"'""".format(server_xml_path)
    jvm_res = ssh_exec(host, user, key, jvm_cmd)

    return {
        "server_xml_path": server_xml_path,
        "heap_lines": heap_res["stdout"].strip() if heap_res["success"] else "",
        "jvm_entries_line": jvm_res["stdout"].strip() if jvm_res["success"] else "",
    }


# -----------------------------------------------------------------------------
# HANDLE SERVER
# -----------------------------------------------------------------------------
def handle_server(group_name, host, profile, global_settings):
    print("\n{0} :: {1}".format(group_name, host))

    per_jvm_info = {}

    def ensure_jvm(j):
        if j not in per_jvm_info:
            per_jvm_info[j] = _new_jvm_report()

    status_result = ssh_exec(
        host,
        profile["ssh_user"],
        profile.get("ssh_key_path"),
        profile["commands"]["status"],
    )

    if not status_result["success"]:
        print("SSH/status failed")
        print(status_result["stderr"])
        cause = (
            status_result.get("stderr") or status_result.get("stdout") or ""
        ).strip() or "unknown"
        return [
            {
                "Patched server": "{0} | {1}".format(host, group_name),
                "status": "Status check failed",
                "cause": cause,
                "ai action": "",
            }
        ]

    configured, started, down = parse_websphere_status(status_result["stdout"])
    initial_configured = list(configured)

    print(
        "Configured JVMs: {0}".format(", ".join(configured) if configured else "None")
    )
    print("Running JVMs   : {0}".format(", ".join(started) if started else "None"))
    print("Stopped JVMs   : {0}".format(", ".join(down) if down else "None"))

    if not down:
        print("All JVMs are running")
        return [_report_row(host, j, "Running", "", "") for j in configured]

    for jvm in down:
        print("\nJVM DOWN: {0}".format(jvm))
        ensure_jvm(jvm)

        match = re.search(r"_(wsvmt\d+)_", jvm)
        if not match:
            print("Cannot determine profile")
            per_jvm_info[jvm]["notes"].append(
                "Cannot determine profile (expected _wsvmt#_ in server name)"
            )
            continue

        profile_suffix = match.group(1)
        profile_dir = "ICI{0}".format(profile_suffix)
        was_home = "/was/wasprofile855/{0}".format(profile_dir)
        log_path = "{0}/logs/{1}/SystemOut.log".format(was_home, jvm)

        timestamp, logs = get_latest_shutdown_context(host, profile, jvm, log_path)
        native_stdout_tail = get_native_stdout_tail(
            host, profile, log_path, lines=200
        )

        shutdown_sections = []
        if logs:
            print("Shutdown at: {0}".format(timestamp))
            shutdown_sections.append(
                "SystemOut.log (context around latest shutdown):\n{0}".format(
                    logs[-8000:]
                )
            )
        if native_stdout_tail:
            shutdown_sections.append(
                "native_stdout.log (last 200 lines; review after SystemOut):\n{0}".format(
                    native_stdout_tail[-4000:]
                )
            )

        if shutdown_sections:
            ai_response = call_ai(
                """
You are a senior WebSphere engineer.

Analyze JVM shutdown.

JVM: {0}

Tasks:
- Exact timestamp
- Trigger (manual / crash / memory / deployment)
- Root cause
- Evidence (log lines)

Logs (prefer SystemOut.log for shutdown markers; use native_stdout.log as supporting JVM/process output):
{1}
""".format(
                    jvm, "\n\n".join(shutdown_sections)
                )
            )

            shutdown_text = format_ai_diagnosis(ai_response)
            per_jvm_info[jvm]["shutdown_ai"] = shutdown_text
            print("\nShutdown Analysis:")
            print(shutdown_text)
        else:
            print("No shutdown logs found")
            per_jvm_info[jvm]["notes"].append(
                "No shutdown context (SystemOut shutdown window or native_stdout tail)"
            )

        start_script = "{0}/bin/startServer.sh".format(was_home)
        start_command = "{0} {1}".format(start_script, jvm)

        retry_count = global_settings.get("retry_count", 1)
        wait_time_seconds = global_settings.get("wait_time_seconds", 60)
        started_successfully = False
        startup_logs = None

        for attempt in range(retry_count):
            print("\nAttempt {0}".format(attempt + 1))

            ssh_exec(
                host, profile["ssh_user"], profile.get("ssh_key_path"), start_command
            )

            time.sleep(wait_time_seconds)

            status_check = ssh_exec(
                host,
                profile["ssh_user"],
                profile.get("ssh_key_path"),
                profile["commands"]["status"],
            )

            _, started_now, _ = parse_websphere_status(status_check["stdout"])

            if jvm in started_now:
                print("Started successfully")
                started_successfully = True
                per_jvm_info[jvm]["recovered"] = True
                break

            if attempt == 0:
                startup_logs = get_startup_failure_context(host, profile, log_path)

        if not started_successfully:
            print("Failed to start")
            per_jvm_info[jvm]["restart_attempted"] = True

            if startup_logs is None:
                startup_logs = get_startup_failure_context(host, profile, log_path)

            if startup_logs:
                generic_l1_context = get_generic_l1_context(
                    host, profile, was_home, jvm, log_path
                )

                heap_details = get_exact_server_xml_and_heap_details(
                    host, profile, was_home, jvm
                )

                diag = call_ai(
                    """
You are a senior WebSphere middleware diagnostics expert.

Analyze ONLY this JVM: {0}
Host: {1}
Profile/Home Path: {2}

Return ONLY valid JSON.
Do not use markdown.
Do not wrap the response in a markdown code block.
Do not add any explanation before or after the JSON.

Provide STRICT JSON response in this format:

{{
  "jvm_that_is_down": "{0}",
  "exact_error": "...",
  "root_cause": "...",
  "severity": "Low|Medium|High|Critical",
  "evidence": "...",
  "explanation": "...",
  "recommended_fix": "...",
  "server_check_path": "...",
  "console_fix_navigation": "...",
  "l1_validation_commands": [
    "command1",
    "command2"
  ],
  "post_validation_recommendation": "...",
  "server_start_command": "{5}",
  "console_fix_navigation": "Servers > Server Types > WebSphere application servers > sharedcl01_wsvmt7_01 > Java and Process Management > Process Definition > Java Virtual Machine"
  "console_start_navigation": "Servers > Server Types > WebSphere application servers > {0} > Select the JVM > Click Start"
}}

Rules for l1_validation_commands:
- Only safe, read-only commands
- No rm, mv, kill, restart, systemctl, chmod, chown, vi, sed -i, startServer, stopServer
- Only use: test, ls, stat, file, netstat, ss, echo, grep, cat, head, tail, find, df, free, ulimit, ps, pwd, id, mount
- Commands must not modify system
- Commands must help validate the suspected root cause
- If exact path is known, include it
- Prefer commands that directly prove the issue, like file existence, permissions, JVM args, memory, disk, port binding, missing mount, exact heap settings, etc.

Important:
- console_fix_navigation is mandatory.
- You must predict the most relevant WebSphere admin console navigation path where the identified issue should be checked or corrected.
- For heap / Xmx / Xms / generic JVM arguments / javaagent issues, console_fix_navigation must be:
  Servers > Server Types > WebSphere application servers > {0} > Java and Process Management > Process Definition > Java Virtual Machine
- For datasource / database connectivity issues, use the appropriate JDBC / datasource console path.
- For SSL / certificate / keystore issues, use the appropriate SSL certificate and key management path.
- For shared library / classpath issues, use the appropriate shared libraries or class loader path.
- Do not guess the server.xml path.
- If the issue is heap related, you must use the exact server.xml path provided below.
- In WebSphere, heap may be defined through initialHeapSize and maximumHeapSize inside jvmEntries, not only through genericJvmArguments.
- If the issue is heap related, provide both:
  1. exact server-side file path to check/update heap
  2. exact WebSphere console navigation path to check/update heap
- If a required file/binary/jar is missing, generate commands that prove it clearly.
- If the issue is configuration related, provide the exact file/path on server in server_check_path.
- The recommendation should be operational, clear, and directly usable by L1/L2 support.
- Mention how to start the JVM after fix from server and from console.

Exact resolved WebSphere config path for this JVM:
Server XML path:
{6}

Exact heap-related entries from server.xml:
{7}

Exact jvmEntries line from server.xml:
{8}

Startup logs (after first failed start: startServer.log first; native_stderr.log only if startServer is inconclusive — see section headers):
{3}

Additional host/runtime diagnostic context:
{4}
""".format(
                        jvm,
                        host,
                        was_home,
                        startup_logs[-8000:],
                        generic_l1_context[-8000:],
                        start_command,
                        heap_details.get("server_xml_path", ""),
                        heap_details.get("heap_lines", ""),
                        heap_details.get("jvm_entries_line", ""),
                    )
                )

                print("\nStartup Failure Analysis:")

                formatted_diag = format_ai_diagnosis(diag)
                print(formatted_diag)

                validation_result = None
                if isinstance(diag, dict):
                    l1_cmds = diag.get("l1_validation_commands", [])
                    if l1_cmds:
                        validation_result = execute_l1_validations(
                            host, profile, l1_cmds
                        )

                validation_text = format_validation_outputs(
                    validation_result, profile=profile, host=host
                )

                if validation_text:
                    print("")
                    print(validation_text)

                ai_text = formatted_diag
                if validation_text:
                    ai_text += "\n\n" + validation_text

                per_jvm_info[jvm]["startup_ai"] = ai_text
            else:
                per_jvm_info[jvm]["notes"].append(
                    "Failed to start; startup logs unavailable for AI analysis"
                )

    final_check = ssh_exec(
        host,
        profile["ssh_user"],
        profile.get("ssh_key_path"),
        profile["commands"]["status"],
    )

    if not final_check["success"]:
        print("\nFinal status recheck failed for host: {0}".format(host))
        rows = []
        for jvm in initial_configured:
            info = per_jvm_info.get(jvm)
            rows.append(
                _report_row(
                    host,
                    jvm,
                    "Unknown",
                    "Final status recheck failed (SSH or empty output)",
                    _jvm_ai_text(info),
                )
            )
        return rows

    final_configured, final_started, final_down = parse_websphere_status(
        final_check["stdout"]
    )

    print("\nFinal node summary for host: {0}".format(host))
    print(
        "Configured JVMs: {0}".format(
            ", ".join(final_configured) if final_configured else "None"
        )
    )
    print(
        "Started JVMs   : {0}".format(
            ", ".join(final_started) if final_started else "None"
        )
    )
    print(
        "Down JVMs      : {0}".format(", ".join(final_down) if final_down else "None")
    )

    if not final_down:
        print("All JVMs are UP")
    else:
        print("Some JVMs are still DOWN")

    rows = []
    for jvm in final_configured:
        running = jvm in final_started
        info = per_jvm_info.get(jvm)
        rows.append(
            _report_row(
                host,
                jvm,
                "Running" if running else "Down",
                _jvm_cause_text(info, running),
                (
                    _jvm_ai_text(info)
                    if (
                        not running
                        or (
                            info and (info.get("shutdown_ai") or info.get("startup_ai"))
                        )
                    )
                    else ""
                ),
            )
        )
    return rows


# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------
def main():
    config = load_config("server_config_websphere.json")

    global_settings = config.get("global_settings", {})
    profiles = config.get("middleware_profiles", {})
    server_groups = config.get("server_groups", {})

    all_rows = []

    for group_name, group in server_groups.items():
        profile_name = group.get("profile")
        profile = profiles.get(profile_name)

        if not profile:
            print("Profile not found: {0}".format(profile_name))
            continue

        for host in group.get("hosts", []):
            rows = handle_server(group_name, host, profile, global_settings)
            if rows:
                all_rows.extend(rows)

    report_path = os.getenv(
        "WEB_AI_REPORT",
        "websphere_monitor_report_{0}.xlsx".format(
            datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        ),
    )
    if write_excel_report(all_rows, report_path):
        print("\nExcel report written: {0}".format(os.path.abspath(report_path)))


# -----------------------------------------------------------------------------
if __name__ == "__main__":
    main()
